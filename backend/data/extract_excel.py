"""
Extract all data from MonsterAdvancer Excel file into SQLite database.
Reads mainDB, attacksDB, armorClassDB, skillsDB, specialsDB, Tables, NamedLists, Glossary.
"""
import json
import os
import sqlite3
import sys

import openpyxl


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "MonsterAdvancer_15.02.xlsm")
DB_PATH = os.path.join(os.path.dirname(__file__), "monsters.db")


def create_schema(conn: sqlite3.Connection):
    c = conn.cursor()

    c.executescript("""
    -- Core monster data from mainDB
    CREATE TABLE IF NOT EXISTS monsters (
        id INTEGER PRIMARY KEY,
        name TEXT NOT NULL,
        size TEXT,
        type TEXT,
        descriptor TEXT,
        hd_count INTEGER,
        hit_dice TEXT,
        initiative TEXT,
        speed TEXT,
        armor_class TEXT,
        base_attack INTEGER,
        grapple TEXT,
        attack TEXT,
        full_attack TEXT,
        space TEXT,
        reach1 TEXT,
        reach2 TEXT,
        reach2_desc TEXT,
        is_reach2_reach_weapon INTEGER,
        reach3 TEXT,
        reach3_desc TEXT,
        special_attacks TEXT,
        special_qualities TEXT,
        skills TEXT,
        feats TEXT,
        all_feats TEXT,
        saves TEXT,
        fort_save_type TEXT,
        ref_save_type TEXT,
        will_save_type TEXT,
        str INTEGER,
        dex INTEGER,
        con INTEGER,
        int INTEGER,
        wis INTEGER,
        cha INTEGER,
        environment TEXT,
        organization TEXT,
        challenge_rating REAL,
        cr_text TEXT,
        treasure TEXT,
        alignment TEXT,
        advancement TEXT,
        max_adv_base_size INTEGER,
        max_adv_next_size INTEGER,
        special_abilities TEXT,
        stat_block TEXT,
        reference TEXT,
        level_adjustment TEXT,
        altname TEXT,
        bonus_feats TEXT,
        bonus_feat_count INTEGER
    );

    -- Attack definitions from attacksDB
    CREATE TABLE IF NOT EXISTS attacks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        monster_id INTEGER NOT NULL,
        group_id INTEGER,
        att_name TEXT,
        att_count INTEGER,
        is_standard INTEGER,
        weapon_nature TEXT,
        att_mode TEXT,
        use_category TEXT,
        dmg_die TEXT,
        crit_range TEXT,
        crit_mult INTEGER,
        str_mult REAL,
        att_roll_enh INTEGER DEFAULT 0,
        dmg_enh INTEGER DEFAULT 0,
        dmg_composite TEXT,
        dmg_text TEXT,
        weapon_text TEXT,
        FOREIGN KEY (monster_id) REFERENCES monsters(id)
    );

    -- AC component breakdown from armorClassDB
    CREATE TABLE IF NOT EXISTS armor_class_components (
        id INTEGER PRIMARY KEY,
        monster_id INTEGER NOT NULL,
        name TEXT,
        base_nat_armor INTEGER DEFAULT 0,
        base_armor INTEGER DEFAULT 0,
        base_armor_description TEXT,
        base_deflection INTEGER DEFAULT 0,
        base_deflection_description TEXT,
        base_shield INTEGER DEFAULT 0,
        base_shield_description TEXT,
        base_dodge INTEGER DEFAULT 0,
        base_dodge_description TEXT,
        FOREIGN KEY (monster_id) REFERENCES monsters(id)
    );

    -- Skill bonuses from skillsDB (stored as JSON per monster)
    CREATE TABLE IF NOT EXISTS monster_skills (
        monster_id INTEGER PRIMARY KEY,
        skills_json TEXT,
        total_skill_points INTEGER,
        skill_count INTEGER,
        FOREIGN KEY (monster_id) REFERENCES monsters(id)
    );

    -- Special abilities text from specialsDB
    CREATE TABLE IF NOT EXISTS special_ability_texts (
        monster_id INTEGER PRIMARY KEY,
        descriptive_text TEXT,
        cosmetic TEXT,
        FOREIGN KEY (monster_id) REFERENCES monsters(id)
    );

    -- Creature type rules from Tables sheet
    CREATE TABLE IF NOT EXISTS type_rules (
        type_name TEXT PRIMARY KEY,
        hd_type INTEGER,
        bab_progression TEXT,
        skill_point_base INTEGER,
        cr_mod INTEGER
    );

    -- Size change stat adjustments from Tables sheet
    CREATE TABLE IF NOT EXISTS size_changes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        size_transition TEXT NOT NULL,
        str_change INTEGER,
        dex_change INTEGER,
        con_change INTEGER,
        nat_armor_change INTEGER,
        ac_change INTEGER,
        attack_change INTEGER
    );

    -- Damage scaling by size from Tables sheet
    CREATE TABLE IF NOT EXISTS damage_scaling (
        old_damage TEXT PRIMARY KEY,
        new_damage TEXT
    );

    -- Feat list from NamedLists
    CREATE TABLE IF NOT EXISTS feats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
    );

    -- Glossary type traits
    CREATE TABLE IF NOT EXISTS type_traits (
        type_name TEXT PRIMARY KEY,
        traits_text TEXT
    );

    -- Named lists (sizes, types, alignments, environments, CR)
    CREATE TABLE IF NOT EXISTS named_lists (
        list_name TEXT NOT NULL,
        value TEXT NOT NULL,
        sort_order INTEGER,
        PRIMARY KEY (list_name, value)
    );

    -- Weapon database (for manufactured weapons - seeded separately)
    CREATE TABLE IF NOT EXISTS weapons (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        category TEXT,  -- Simple, Martial, Exotic
        subcategory TEXT,  -- Unarmed, Light Melee, One-Handed Melee, Two-Handed Melee, Ranged
        cost TEXT,
        dmg_small TEXT,
        dmg_medium TEXT,
        critical TEXT,
        range_increment TEXT,
        weight REAL,
        damage_type TEXT,  -- Bludgeoning, Piercing, Slashing
        special TEXT
    );

    -- Armor database (seeded separately)
    CREATE TABLE IF NOT EXISTS armor (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        category TEXT,  -- Light, Medium, Heavy, Shield
        ac_bonus INTEGER,
        max_dex INTEGER,
        check_penalty INTEGER,
        spell_failure INTEGER,
        speed_30 TEXT,
        speed_20 TEXT,
        weight REAL,
        cost TEXT
    );

    -- Class progression table (seeded separately)
    CREATE TABLE IF NOT EXISTS class_progression (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_name TEXT NOT NULL,
        level INTEGER NOT NULL,
        bab INTEGER,
        fort_save INTEGER,
        ref_save INTEGER,
        will_save INTEGER,
        hd_type INTEGER,
        skill_points_per_level INTEGER,
        features TEXT,
        UNIQUE(class_name, level)
    );

    -- Class definitions
    CREATE TABLE IF NOT EXISTS classes (
        name TEXT PRIMARY KEY,
        category TEXT,  -- NPC, Base
        hd_type INTEGER,
        bab_progression TEXT,  -- good, average, poor
        fort_progression TEXT,  -- good, poor
        ref_progression TEXT,
        will_progression TEXT,
        skill_points_base INTEGER,
        weapon_proficiency TEXT,
        armor_proficiency TEXT,
        description TEXT
    );

    CREATE INDEX IF NOT EXISTS idx_attacks_monster ON attacks(monster_id);
    CREATE INDEX IF NOT EXISTS idx_monsters_type ON monsters(type);
    CREATE INDEX IF NOT EXISTS idx_monsters_cr ON monsters(challenge_rating);
    CREATE INDEX IF NOT EXISTS idx_monsters_size ON monsters(size);
    """)
    conn.commit()


def safe_int(val, default=None):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_float(val, default=None):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def parse_cr(val):
    """
    Parse a CR value that may contain parenthetical variants or fraction strings.
    Returns (numeric_float, display_text).
    Examples:
      "5 (noble 8)"                      -> (5.0, "5 (noble 8)")
      "4 (5 with irresistible dance)"    -> (4.0, "4 (5 with irresistible dance)")
      "2 (without pipes) or 4 ..."       -> (2.0, "2 (without pipes) or 4 ...")
      "4 (normal);\n6 (pyro- or cryo-)" -> (4.0, "4 (normal); 6 (pyro- or cryo-)")
      "1/2"                              -> (0.5, "1/2")
      5.0                                -> (5.0, "5")
    """
    import re
    if val is None:
        return None, None
    # Already a number
    if isinstance(val, (int, float)):
        num = float(val)
        text = str(int(num)) if num == int(num) else str(num)
        return num, text
    s = str(val).strip()
    s = re.sub(r';\s*\n\s*', '; ', s)   # ";\n" -> "; "
    s = re.sub(r'\n+', '; ', s)         # bare newlines -> "; "
    # Grab first token (before space, parenthesis, semicolon, or "or")
    first = re.split(r'[\s(;]', s)[0]
    # Handle fraction like "1/2"
    if '/' in first:
        parts = first.split('/')
        try:
            num = float(parts[0]) / float(parts[1])
            return num, s
        except (ValueError, ZeroDivisionError):
            pass
    # Handle plain number
    try:
        num = float(first)
        return num, s
    except ValueError:
        pass
    return None, s


def safe_str(val):
    if val is None:
        return None
    return str(val).strip()


def extract_monsters(wb, conn):
    ws = wb["mainDB"]
    c = conn.cursor()
    count = 0

    for row in range(3, ws.max_row + 1):
        monster_id = safe_int(ws.cell(row=row, column=1).value)
        name = safe_str(ws.cell(row=row, column=2).value)
        if not name or name == "Select a monster":
            continue

        c.execute("""
            INSERT OR REPLACE INTO monsters (
                id, name, size, type, descriptor, hd_count, hit_dice,
                initiative, speed, armor_class, base_attack, grapple,
                attack, full_attack, space, reach1, reach2, reach2_desc,
                is_reach2_reach_weapon, reach3, reach3_desc,
                special_attacks, special_qualities, skills, feats, all_feats,
                saves, fort_save_type, ref_save_type, will_save_type,
                str, dex, con, int, wis, cha,
                environment, organization, challenge_rating, cr_text, treasure,
                alignment, advancement, max_adv_base_size, max_adv_next_size,
                special_abilities, stat_block, reference, level_adjustment,
                altname, bonus_feats, bonus_feat_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            monster_id,
            name,
            safe_str(ws.cell(row=row, column=3).value),   # size
            safe_str(ws.cell(row=row, column=4).value),   # type
            safe_str(ws.cell(row=row, column=5).value),   # descriptor
            safe_int(ws.cell(row=row, column=6).value),   # hd_count
            safe_str(ws.cell(row=row, column=7).value),   # hit_dice
            safe_str(ws.cell(row=row, column=8).value),   # initiative
            safe_str(ws.cell(row=row, column=9).value),   # speed
            safe_str(ws.cell(row=row, column=10).value),  # armor_class
            safe_int(ws.cell(row=row, column=11).value),  # base_attack
            safe_str(ws.cell(row=row, column=12).value),  # grapple
            safe_str(ws.cell(row=row, column=13).value),  # attack
            safe_str(ws.cell(row=row, column=14).value),  # full_attack
            safe_str(ws.cell(row=row, column=15).value),  # space
            safe_str(ws.cell(row=row, column=16).value),  # reach1
            safe_str(ws.cell(row=row, column=17).value),  # reach2
            safe_str(ws.cell(row=row, column=18).value),  # reach2_desc
            safe_int(ws.cell(row=row, column=19).value),  # is_reach2_reach_weapon
            safe_str(ws.cell(row=row, column=20).value),  # reach3
            safe_str(ws.cell(row=row, column=21).value),  # reach3_desc
            safe_str(ws.cell(row=row, column=22).value),  # special_attacks
            safe_str(ws.cell(row=row, column=23).value),  # special_qualities
            safe_str(ws.cell(row=row, column=24).value),  # skills
            safe_str(ws.cell(row=row, column=25).value),  # feats
            safe_str(ws.cell(row=row, column=26).value),  # all_feats
            safe_str(ws.cell(row=row, column=27).value),  # saves
            safe_str(ws.cell(row=row, column=28).value),  # fort_save_type
            safe_str(ws.cell(row=row, column=29).value),  # ref_save_type
            safe_str(ws.cell(row=row, column=30).value),  # will_save_type
            safe_int(ws.cell(row=row, column=31).value),  # str
            safe_int(ws.cell(row=row, column=32).value),  # dex
            safe_int(ws.cell(row=row, column=33).value),  # con
            safe_int(ws.cell(row=row, column=34).value),  # int
            safe_int(ws.cell(row=row, column=35).value),  # wis
            safe_int(ws.cell(row=row, column=36).value),  # cha
            safe_str(ws.cell(row=row, column=37).value),  # environment
            safe_str(ws.cell(row=row, column=38).value),  # organization
            *parse_cr(ws.cell(row=row, column=39).value),  # challenge_rating, cr_text
            safe_str(ws.cell(row=row, column=40).value),  # treasure
            safe_str(ws.cell(row=row, column=41).value),  # alignment
            safe_str(ws.cell(row=row, column=42).value),  # advancement
            safe_int(ws.cell(row=row, column=43).value),  # max_adv_base_size
            safe_int(ws.cell(row=row, column=44).value),  # max_adv_next_size
            safe_str(ws.cell(row=row, column=45).value),  # special_abilities
            safe_str(ws.cell(row=row, column=46).value),  # stat_block
            safe_str(ws.cell(row=row, column=47).value),  # reference
            safe_str(ws.cell(row=row, column=48).value),  # level_adjustment
            safe_str(ws.cell(row=row, column=49).value),  # altname
            safe_str(ws.cell(row=row, column=50).value),  # bonus_feats
            safe_int(ws.cell(row=row, column=51).value),  # bonus_feat_count
        ))
        count += 1

    # Animals and Vermin are always neutral; Excel omits it for most entries
    c.execute("""
        UPDATE monsters SET alignment = 'Always neutral'
        WHERE type IN ('Animal', 'Vermin') AND (alignment IS NULL OR alignment = '')
    """)
    fixed = c.rowcount
    if fixed:
        print(f"  Defaulted {fixed} Animal/Vermin alignments to 'Always neutral'")

    conn.commit()
    print(f"  Extracted {count} monsters")


def extract_attacks(wb, conn):
    ws = wb["attacksDB"]
    c = conn.cursor()
    count = 0

    for row in range(8, ws.max_row + 1):
        monster_id = safe_int(ws.cell(row=row, column=1).value)
        att_name = safe_str(ws.cell(row=row, column=3).value)
        if monster_id is None or att_name is None:
            continue

        c.execute("""
            INSERT INTO attacks (
                monster_id, group_id, att_name, att_count, is_standard,
                weapon_nature, att_mode, use_category, dmg_die,
                crit_range, crit_mult, str_mult, att_roll_enh, dmg_enh,
                dmg_composite, dmg_text, weapon_text
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            monster_id,
            safe_int(ws.cell(row=row, column=2).value),
            att_name,
            safe_int(ws.cell(row=row, column=4).value, 1),
            1 if ws.cell(row=row, column=5).value else 0,
            safe_str(ws.cell(row=row, column=6).value),
            safe_str(ws.cell(row=row, column=7).value),
            safe_str(ws.cell(row=row, column=8).value),
            safe_str(ws.cell(row=row, column=9).value),
            safe_str(ws.cell(row=row, column=10).value),
            safe_int(ws.cell(row=row, column=11).value, 2),
            safe_float(ws.cell(row=row, column=12).value, 1.0),
            safe_int(ws.cell(row=row, column=13).value, 0),
            safe_int(ws.cell(row=row, column=14).value, 0),
            safe_str(ws.cell(row=row, column=15).value),
            safe_str(ws.cell(row=row, column=16).value),
            safe_str(ws.cell(row=row, column=17).value),
        ))
        count += 1

    conn.commit()
    print(f"  Extracted {count} attack definitions")


def extract_armor_class(wb, conn):
    ws = wb["armorClassDB"]
    c = conn.cursor()
    count = 0

    for row in range(3, ws.max_row + 1):
        monster_id = safe_int(ws.cell(row=row, column=1).value)
        name = safe_str(ws.cell(row=row, column=2).value)
        if monster_id is None or not name:
            continue

        c.execute("""
            INSERT OR REPLACE INTO armor_class_components (
                id, monster_id, name, base_nat_armor, base_armor,
                base_armor_description, base_deflection, base_deflection_description,
                base_shield, base_shield_description, base_dodge, base_dodge_description
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            monster_id,
            monster_id,
            name,
            safe_int(ws.cell(row=row, column=3).value, 0),
            safe_int(ws.cell(row=row, column=4).value, 0),
            safe_str(ws.cell(row=row, column=5).value),
            safe_int(ws.cell(row=row, column=6).value, 0),
            safe_str(ws.cell(row=row, column=7).value),
            safe_int(ws.cell(row=row, column=8).value, 0),
            safe_str(ws.cell(row=row, column=9).value),
            safe_int(ws.cell(row=row, column=10).value, 0),
            safe_str(ws.cell(row=row, column=11).value),
        ))
        count += 1

    conn.commit()
    print(f"  Extracted {count} AC component records")


def extract_skills(wb, conn):
    ws = wb["skillsDB"]
    c = conn.cursor()
    count = 0

    headers = []
    for col in range(2, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers.append((col, str(h)))

    for row in range(2, ws.max_row + 1):
        monster_id = safe_int(ws.cell(row=row, column=1).value)
        if monster_id is None:
            continue

        skills = {}
        for col, skill_name in headers:
            val = ws.cell(row=row, column=col).value
            if val is not None and val != "" and val != 0:
                skills[skill_name] = val

        if not skills:
            continue

        # Get total and count from last 2 columns
        total_sp = safe_int(ws.cell(row=row, column=ws.max_column - 1).value, 0)
        skill_count = safe_int(ws.cell(row=row, column=ws.max_column).value, 0)

        c.execute("""
            INSERT OR REPLACE INTO monster_skills (
                monster_id, skills_json, total_skill_points, skill_count
            ) VALUES (?, ?, ?, ?)
        """, (
            monster_id,
            json.dumps(skills),
            total_sp,
            skill_count,
        ))
        count += 1

    conn.commit()
    print(f"  Extracted {count} skill records")


def extract_specials(wb, conn):
    ws = wb["specialsDB"]
    c = conn.cursor()
    count = 0

    for row in range(2, ws.max_row + 1):
        monster_id = safe_int(ws.cell(row=row, column=1).value)
        text = safe_str(ws.cell(row=row, column=2).value)
        cosmetic = safe_str(ws.cell(row=row, column=3).value)
        if monster_id is None:
            continue

        c.execute("""
            INSERT OR REPLACE INTO special_ability_texts (
                monster_id, descriptive_text, cosmetic
            ) VALUES (?, ?, ?)
        """, (monster_id, text, cosmetic))
        count += 1

    conn.commit()
    print(f"  Extracted {count} special ability texts")


def extract_type_rules(wb, conn):
    ws = wb["Tables"]
    c = conn.cursor()
    count = 0

    for row in range(2, 17):
        type_name = safe_str(ws.cell(row=row, column=9).value)
        if not type_name:
            continue

        c.execute("""
            INSERT OR REPLACE INTO type_rules (
                type_name, hd_type, bab_progression, skill_point_base, cr_mod
            ) VALUES (?, ?, ?, ?, ?)
        """, (
            type_name,
            safe_int(ws.cell(row=row, column=10).value),
            safe_str(ws.cell(row=row, column=11).value),
            safe_int(ws.cell(row=row, column=12).value),
            safe_int(ws.cell(row=row, column=13).value),
        ))
        count += 1

    conn.commit()
    print(f"  Extracted {count} type rules")


def extract_size_changes(wb, conn):
    ws = wb["Tables"]
    c = conn.cursor()
    count = 0

    for row in range(3, 11):
        transition = safe_str(ws.cell(row=row, column=1).value)
        if not transition or "Size" in transition:
            continue

        c.execute("""
            INSERT INTO size_changes (
                size_transition, str_change, dex_change, con_change,
                nat_armor_change, ac_change, attack_change
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            transition,
            safe_int(ws.cell(row=row, column=2).value, 0),
            safe_int(ws.cell(row=row, column=3).value, 0),
            safe_int(ws.cell(row=row, column=4).value, 0),
            safe_int(ws.cell(row=row, column=5).value, 0),
            safe_int(ws.cell(row=row, column=6).value, 0),
            safe_int(ws.cell(row=row, column=7).value, 0),
        ))
        count += 1

    conn.commit()
    print(f"  Extracted {count} size change rules")


def extract_damage_scaling(wb, conn):
    ws = wb["Tables"]
    c = conn.cursor()
    count = 0

    for row in range(14, 35):
        old_dmg = safe_str(ws.cell(row=row, column=1).value)
        new_dmg = safe_str(ws.cell(row=row, column=2).value)
        if not old_dmg or old_dmg == "Old Damage" or not new_dmg:
            continue

        c.execute("""
            INSERT OR REPLACE INTO damage_scaling (old_damage, new_damage)
            VALUES (?, ?)
        """, (old_dmg.strip(), new_dmg.strip()))
        count += 1

    conn.commit()
    print(f"  Extracted {count} damage scaling entries")


def extract_feats(wb, conn):
    ws = wb["NamedLists"]
    c = conn.cursor()
    count = 0

    for row in range(1, ws.max_row + 1):
        feat_name = safe_str(ws.cell(row=row, column=3).value)
        if not feat_name or feat_name == "Feat Name":
            continue

        c.execute("""
            INSERT OR IGNORE INTO feats (name) VALUES (?)
        """, (feat_name,))
        count += 1

    conn.commit()
    print(f"  Extracted {count} feats")


def extract_glossary(wb, conn):
    ws = wb["Glossary"]
    c = conn.cursor()
    count = 0

    for row in range(2, ws.max_row + 1):
        type_name = safe_str(ws.cell(row=row, column=1).value)
        traits = safe_str(ws.cell(row=row, column=2).value)
        if not type_name:
            continue

        c.execute("""
            INSERT OR REPLACE INTO type_traits (type_name, traits_text)
            VALUES (?, ?)
        """, (type_name, traits))
        count += 1

    conn.commit()
    print(f"  Extracted {count} type traits")


def extract_named_lists(wb, conn):
    ws = wb["NamedLists"]
    c = conn.cursor()

    list_configs = [
        (1, "sizes"),
        (5, "challenge_ratings"),
        (7, "creature_types"),
        (9, "environments"),
        (10, "size_filter"),
        (11, "alignments"),
    ]

    total = 0
    for col, list_name in list_configs:
        for row in range(1, ws.max_row + 1):
            val = safe_str(ws.cell(row=row, column=col).value)
            if val:
                c.execute("""
                    INSERT OR IGNORE INTO named_lists (list_name, value, sort_order)
                    VALUES (?, ?, ?)
                """, (list_name, val, row))
                total += 1

    conn.commit()
    print(f"  Extracted {total} named list entries")


def main():
    print("=" * 60)
    print("MONSTER ADVANCER - Excel to SQLite Extraction")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    # Remove existing DB
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"Removed existing database: {DB_PATH}")

    print(f"\nLoading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Worksheets: {wb.sheetnames}")

    conn = sqlite3.connect(DB_PATH)

    print("\nCreating database schema...")
    create_schema(conn)

    print("\nExtracting data:")
    print("  [1/9] Monsters (mainDB)...")
    extract_monsters(wb, conn)

    print("  [2/9] Attacks (attacksDB)...")
    extract_attacks(wb, conn)

    print("  [3/9] Armor Class (armorClassDB)...")
    extract_armor_class(wb, conn)

    print("  [4/9] Skills (skillsDB)...")
    extract_skills(wb, conn)

    print("  [5/9] Special Abilities (specialsDB)...")
    extract_specials(wb, conn)

    print("  [6/9] Type Rules (Tables)...")
    extract_type_rules(wb, conn)

    print("  [7/9] Size Changes & Damage Scaling (Tables)...")
    extract_size_changes(wb, conn)
    extract_damage_scaling(wb, conn)

    print("  [8/9] Feats (NamedLists)...")
    extract_feats(wb, conn)

    print("  [9/9] Glossary & Named Lists...")
    extract_glossary(wb, conn)
    extract_named_lists(wb, conn)

    # Print summary
    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY")
    print("=" * 60)
    cursor = conn.cursor()
    tables = [
        "monsters", "attacks", "armor_class_components", "monster_skills",
        "special_ability_texts", "type_rules", "size_changes", "damage_scaling",
        "feats", "type_traits", "named_lists",
    ]
    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM {table}")
        count = cursor.fetchone()[0]
        print(f"  {table}: {count} rows")

    db_size = os.path.getsize(DB_PATH) / 1024
    print(f"\nDatabase size: {db_size:.1f} KB")
    print(f"Database path: {DB_PATH}")

    conn.close()
    wb.close()
    print("\nExtraction complete!")


if __name__ == "__main__":
    main()
